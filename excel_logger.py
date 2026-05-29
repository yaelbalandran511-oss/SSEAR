"""
Registro de evaluaciones en archivo Excel para SSEAR.
"""

import os
import logging
from datetime import datetime
from typing import Dict

import config
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

EXCEL_FILE = os.path.join(config.LOGS_DIR, 'evaluation_results.xlsx')
HEADERS = [
    'Fecha y hora',
    'Pregunta',
    'Respuesta de referencia',
    'Respuesta del estudiante',
    'Similitud semántica',
    'Similitud léxica',
    'Calificación general',
    'Términos encontrados',
    'Términos faltantes',
    'Términos extra',
    'Retroalimentación generada'
]


def _normalize_list(values):
    if not values:
        return ''
    return ', '.join(str(v) for v in values)


def _ensure_headers(ws):
    ws.append(HEADERS)
    for index, header in enumerate(HEADERS, start=1):
        width = max(len(header) + 2, 18)
        ws.column_dimensions[get_column_letter(index)].width = width


def append_evaluation_to_excel(question: str, reference_answer: str, student_answer: str,
                               evaluation_result: Dict, feedback: Dict) -> None:
    """Agrega una fila con los resultados de evaluación al archivo Excel."""
    try:
        os.makedirs(config.LOGS_DIR, exist_ok=True)

        if os.path.exists(EXCEL_FILE):
            workbook = load_workbook(EXCEL_FILE)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Evaluaciones'
            _ensure_headers(worksheet)

        scores = evaluation_result.get('scores', {})
        metadata = evaluation_result.get('metadata', {})

        row = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            question,
            reference_answer,
            student_answer,
            scores.get('semantic', 0),
            scores.get('lexical', 0),
            scores.get('percentage', scores.get('overall', 0)),
            _normalize_list(metadata.get('matched_terms')),
            _normalize_list(metadata.get('missing_terms')),
            _normalize_list(metadata.get('extra_terms')),
            _normalize_list(feedback.get('specific_feedback') or [feedback.get('summary', {}).get('message', '')])
        ]

        worksheet.append(row)
        workbook.save(EXCEL_FILE)
        logger.info(f"Evaluación registrada en Excel: {EXCEL_FILE}")
    except Exception as exc:
        logger.warning(f"No se pudo registrar la evaluación en Excel: {exc}", exc_info=True)
