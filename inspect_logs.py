import openpyxl
from pathlib import Path
path = Path('logs') / 'evaluation_results.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print('Headers:', headers)
print('First 5 rows:')
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    print(row)
print('Total rows:', max(0, ws.max_row - 1))
